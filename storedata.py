import shutil
from io import StringIO
import json
import sys
import os

import time

import pymysql
from django.utils.six import BytesIO
pwd = os.path.dirname(os.path.realpath(__file__))
# 获取项目名的目录(因为我的当前文件是在项目名下的文件夹下的文件.所以是../)
from openpyxl import load_workbook

sys.path.append(pwd+"../")
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "软工.settings")
import requests
# 获取当前文件的目录
from rest_framework.parsers import JSONParser
from rest_framework.renderers import JSONRenderer


import django
django.setup()
import data_spider
from player_data.persons.models import Player, Team, Record ,Match_teamsummary, Match, Match_player, Score, Schedule
from player_data.persons.serializers import PlayerSerializer, TeamSerializer, RecordSerializer, CareerSerializer,MatchSerializer,Match_playerSerializer,Match_teamsummarySerializer,ScoreSerializer,ScheduleSerializer
from django.core.files import File
#接下来就可以使用model了

def history_in_database(macth_id):

    if Match.objects.filter(id=macth_id).count()==0:
        return 0
    else:
        return 1


def delete_files() :
    os.chdir("./history_games(date)")
    fileList = list(os.listdir())
    for file in fileList:
        if os.path.isfile(file):
            os.remove(file)
            print("delete successfully")
        else:
            shutil.rmtree(file)
#添加球员信息
# with open('./active_players(1).json') as f:
#     active_players = json.load(f)
# active_players = active_players['active_players']
#
# for player in active_players:
#     try:
#         if (player.get('personId')!=None)&(Player.objects.filter(personId=player.get('personId')).count()==0):
#             image_path="player/"+player.get('personId')+".png"
#             try:
#                 image=open(image_path,'rb')
#             except:
#                 image=open("player/0000.png",'rb')
#             image=File(image)
#             content = JSONRenderer().render(player)
#             stream = BytesIO(content)
#             data = JSONParser().parse(stream)
#             serializer = PlayerSerializer(data=data)
#             if serializer.is_valid():
#                 serializer.save()
#                 playerob=Player.objects.get(personId=player.get('personId'))
#                 playerob.playerLogo=image
#                 playerob.save()
#             else:
#                 print(serializer.errors)
#     except Exception:
#         print(Exception.with_traceback())
# #添加球队
# with open('./team_config.json') as f:
#     team_list= json.load(f)
# team_list=team_list['teams']['config']
#
# for team in team_list:
#     if team.get('ttsName')!=None:
#         image_path="team/"+team.get('tricode')+"_logo.svg"
#         image=open(image_path)
#         image=File(image)
#
#         content = JSONRenderer().render(team)
#         stream = BytesIO(content)
#         data = JSONParser().parse(stream)
#         #serializer = TeamSerializer(data=data)
#         if 1:
#             #serializer.save()
#             Teamob=Team.objects.get(teamId=team['teamId'])
#             Teamob.teamLogo=image
#             Teamob.save()
#         else:
#             print(serializer.errors)

# glg = data_spider.GetLiveGames()
# # glg.get_scores()
# # glg.get_active_players()
# # glg.get_schedule()
# # glg.get_player_profile()
# with open('./active_players(1).json') as f:
#     active_players = json.load(f)
# active_players = active_players['active_players']
# # profile = {}
# for player in active_players:
#     if (player.get('personId')!=None):
#         player_profile = requests.get(glg.player_profile.format(player['personId'])).json()
#     # profile[player['personId']] = player_profile
#         player_profile = json.dumps(player_profile, sort_keys=True, indent=8, separators=(',', ':'))
#         career_total=player_profile['league']['standard']['stats']['careerSummary']


import os


# file_dir="./team_info"
# team_index=0
# index=0
# for root, dirs, files in os.walk(file_dir):
#     print(files) #当前路径下所有非目录子文件
#     for file in files:
#         with open(root+"/"+file) as f:
#             team_info = json.load(f)
#         print(team_info)
#         team_info=team_info[file.split('.')[0]]
#         team_info['球队名']=file
#         info=team_info['技术统计']
#         for item in info:
#             item_info=info[str(item)]
#             item_info['序号']=index
#             index=index+1
#             if Record.objects.filter(序号=index).count()!=0:
#                 continue
#             content = JSONRenderer().render(item_info)
#             stream = BytesIO(content)
#             data = JSONParser().parse(stream)
#             serializer = RecordSerializer(data=data)
#             if serializer.is_valid():
#                 serializer.save()
#             else:
#                 print(serializer.errors)
#
#         team_info['场均助攻']=str(team_index*5+0)
#         team_info['场均失分']=str(team_index*5+1)
#         team_info['场均失误']=str(team_index*5+2)
#         team_info['场均得分']=str(team_index*5+3)
#         team_info['场均篮板']=str(team_index*5+4)
#
        # content = JSONRenderer().render(team_info)
        # stream = BytesIO(content)
        # data = JSONParser().parse(stream)
        # serializer = TeamSerializer(data=data)
        #
        # if serializer.is_valid():
        #     serializer.save()
        # else:
        #     print(serializer.errors)
#         team_index=team_index+1


# db = pymysql.connect("114.116.156.240", "root", "Buaa2019!", "app", charset='utf8')
# cursor = db.cursor()
import xlrd, xlwt


def store_history_game():
    file_dir2="./history_games(date)"
    dir_list = os.listdir(file_dir2) # 列出文件夹下所有的目录和文件
    i=0
    for i in range(len(dir_list)): # 遍历文件夹内所有子文件夹
        # 判断是否为已存文件
        num=Match.objects.filter(id=dir_list[i].split('-')[3]).count()
        if  (num==0) | ((dir_list[i].split('-')[0]+'-'+dir_list[i].split('-')[1]+'-'+dir_list[i].split('-')[2])==time.strftime("%Y-%m-%d")):
            # print(dir_list[i])
        # 若未存，则开始存子文件夹内文件
            if num==1:
                matchob=Match.objects.get(id=dir_list[i].split('-')[3])
                matchob.delete()
            print(dir_list[i], '---------dont exist, start to store')
            new_path = os.path.join(file_dir2, dir_list[i])
            file_list = os.listdir(new_path)

            real_path1 = os.path.join(new_path, file_list[0])
            real_path2 = os.path.join(new_path, file_list[1])
            real_path3 = os.path.join(new_path, file_list[2])

            workbook1 =  load_workbook(real_path1)
            sheet1 = workbook1.get_sheet_by_name("Sheet1")
            workbook2 =  load_workbook(real_path2)
            sheet2 = workbook2.get_sheet_by_name("Sheet1")
            workbook3 =  load_workbook(real_path3)
            sheet3 = workbook3.get_sheet_by_name("Sheet1")

            if sheet1.max_row > 10: # 判断是否为空表

                # 存Match内容，区分是否有加时
                match = Match(id=dir_list[i].split('-')[3],
                              日期=dir_list[i].split('-')[0]+'-'+dir_list[i].split('-')[1]+'-'+dir_list[i].split('-')[2],
                              主场球队中文名=sheet3.cell_value(3, 1),
                              主场第一节=sheet3.cell_value(3, 2),
                              主场第二节=sheet3.cell_value(3, 3),
                              主场第三节=sheet3.cell_value(3, 4),
                              主场第四节=sheet3.cell_value(3, 5),
                              主场总分=sheet3.cell_value(3, 6), # 没有加时的比赛总分
                              主场加时一='0',
                              主场加时二='0',
                              主场加时三='0',
                              主场加时四='0',
                              客场加时一='0',
                              客场加时二='0',
                              客场加时三='0',
                              客场加时四='0',
                              客场球队中文名=sheet3.cell_value(2, 1),
                              客场第一节=sheet3.cell_value(2, 2),
                              客场第二节=sheet3.cell_value(2, 3),
                              客场第三节=sheet3.cell_value(2, 4),
                              客场第四节=sheet3.cell_value(2, 5),
                              客场总分=sheet3.cell_value(2, 6), # 没有加时的比赛总分
                              )
                if sheet3.cell_value(1, 6)=='加时一':
                    match.主场加时一=sheet3.cell_value(3, 6)
                    match.客场加时一=sheet3.cell_value(2, 6)
                    if sheet3.cell_value(1, 7)=='加时二':
                        match.主场加时二=sheet3.cell_value(3, 7)
                        match.客场加时二=sheet3.cell_value(2, 7)
                        if sheet3.cell_value(1, 8)=='加时三':
                            match.主场加时三=sheet3.cell_value(3, 8)
                            match.客场加时三=sheet3.cell_value(2, 8)
                            if sheet3.cell_value(1, 9)=='加时四':
                                match.主场加时四=sheet3.cell_value(3, 9)
                                match.客场加时四=sheet3.cell_value(2, 9)
                                match.主场总分=sheet3.cell_value(3, 10)
                                match.客场总分=sheet3.cell_value(2, 10)
                            else:
                                match.主场总分=sheet3.cell_value(3, 9)
                                match.客场总分=sheet3.cell_value(2, 9)
                        else:
                            match.主场总分=sheet3.cell_value(3, 8)
                            match.客场总分=sheet3.cell_value(2, 8)
                    else:
                        match.主场总分=sheet3.cell_value(3, 7)
                        match.客场总分=sheet3.cell_value(2, 7)
                match.save()
                # 存球队汇总，区分主客场
                # 客场
                away_summary = Match_teamsummary(主客场=sheet3.cell_value(2, 1), home_away=1,
                                                 投篮=sheet1.cell_value(sheet1.nrows-2, 4), 三分=sheet1.cell_value(sheet1.nrows-2, 5), 罚球=sheet1.cell_value(sheet1.nrows-2, 6),
                                                 前场=sheet1.cell_value(sheet1.nrows-2, 7), 后场=sheet1.cell_value(sheet1.nrows-2, 8), 篮板=sheet1.cell_value(sheet1.nrows-2, 9),
                                                 助攻=sheet1.cell_value(sheet1.nrows-2, 10), 犯规=sheet1.cell_value(sheet1.nrows-2, 11), 抢断=sheet1.cell_value(sheet1.nrows-2, 12),
                                                 失误=sheet1.cell_value(sheet1.nrows-2, 13), 封盖=sheet1.cell_value(sheet1.nrows-2, 14), 得分=sheet1.cell_value(sheet1.nrows-2, 15),
                                                 投篮命中率=sheet1.cell_value(sheet1.nrows-1, 4), 三分命中率=sheet1.cell_value(sheet1.nrows-1, 5), 罚球命中率=sheet1.cell_value(sheet1.nrows-1, 6),
                                                 比赛id=Match.objects.get(id=dir_list[i].split('-')[3]))

                away_summary.save()
                # 主场
                home_summary = Match_teamsummary(主客场=sheet3.cell_value(3, 1), home_away=2,
                                                 投篮=sheet2.cell_value(sheet2.nrows - 2, 4),
                                                 三分=sheet2.cell_value(sheet2.nrows - 2, 5),
                                                 罚球=sheet2.cell_value(sheet2.nrows - 2, 6),
                                                 前场=sheet2.cell_value(sheet2.nrows - 2, 7),
                                                 后场=sheet2.cell_value(sheet2.nrows - 2, 8),
                                                 篮板=sheet2.cell_value(sheet2.nrows - 2, 9),
                                                 助攻=sheet2.cell_value(sheet2.nrows - 2, 10),
                                                 犯规=sheet2.cell_value(sheet2.nrows - 2, 11),
                                                 抢断=sheet2.cell_value(sheet2.nrows - 2, 12),
                                                 失误=sheet2.cell_value(sheet2.nrows - 2, 13),
                                                 封盖=sheet2.cell_value(sheet2.nrows - 2, 14),
                                                 得分=sheet2.cell_value(sheet2.nrows - 2, 15),
                                                 投篮命中率=sheet2.cell_value(sheet2.nrows - 1, 4),
                                                 三分命中率=sheet2.cell_value(sheet2.nrows - 1, 5),
                                                 罚球命中率=sheet2.cell_value(sheet2.nrows - 1, 6),
                                                 比赛id=Match.objects.get(id=dir_list[i].split('-')[3]))

                home_summary.save()
                # 存球员信息，区分主客场，区分首发替补
                for j in range(2, 7):  # 首发 客场

                    match_player = Match_player(类型='首发', 主客场=sheet3.cell_value(2, 1), 球员名=sheet1.cell_value(j, 1),
                                                位置=sheet1.cell_value(j, 2), 时间=sheet1.cell_value(j, 3),
                                                投篮=sheet1.cell_value(j, 4), 三分=sheet1.cell_value(j, 5),
                                                罚球=sheet1.cell_value(j, 6),
                                                前场=sheet1.cell_value(j, 7), 后场=sheet1.cell_value(j, 8),
                                                篮板=sheet1.cell_value(j, 9),
                                                助攻=sheet1.cell_value(j, 10), 犯规=sheet1.cell_value(j, 11),
                                                抢断=sheet1.cell_value(j, 12),
                                                失误=sheet1.cell_value(j, 13), 封盖=sheet1.cell_value(j, 14),
                                                得分=sheet1.cell_value(j, 15),
                                                正负=sheet1.cell_value(j, 16),
                                                比赛id=Match.objects.get(id=dir_list[i].split('-')[3]))

                    match_player.save()

                for k in range (8, sheet1.nrows-2): # 替补

                    match_player = Match_player(类型='替补', 主客场=sheet3.cell_value(2, 1), 球员名=sheet1.cell_value(k, 1), 位置=sheet1.cell_value(k, 2), 时间=sheet1.cell_value(k, 3),
                                                投篮=sheet1.cell_value(k, 4), 三分=sheet1.cell_value(k, 5), 罚球=sheet1.cell_value(k, 6),
                                                前场=sheet1.cell_value(k, 7), 后场=sheet1.cell_value(k, 8), 篮板=sheet1.cell_value(k, 9),
                                                助攻=sheet1.cell_value(k, 10), 犯规=sheet1.cell_value(k, 11), 抢断=sheet1.cell_value(k, 12),
                                                失误=sheet1.cell_value(k, 13), 封盖=sheet1.cell_value(k, 14), 得分=sheet1.cell_value(k, 15),
                                                正负=sheet1.cell_value(k, 16),
                                                比赛id=Match.objects.get(id=dir_list[i].split('-')[3]))
                    match_player.save()

                for j in range (2, 7): # 首发 主场

                    match_player = Match_player(类型='首发', 主客场=sheet3.cell_value(3, 1), 球员名=sheet2.cell_value(j, 1), 位置=sheet2.cell_value(j, 2), 时间=sheet2.cell_value(j, 3),
                                                投篮=sheet2.cell_value(j, 4), 三分=sheet2.cell_value(j, 5), 罚球=sheet2.cell_value(j, 6),
                                                前场=sheet2.cell_value(j, 7), 后场=sheet2.cell_value(j, 8), 篮板=sheet2.cell_value(j, 9),
                                                助攻=sheet2.cell_value(j, 10), 犯规=sheet2.cell_value(j, 11), 抢断=sheet2.cell_value(j, 12),
                                                失误=sheet2.cell_value(j, 13), 封盖=sheet2.cell_value(j, 14), 得分=sheet2.cell_value(j, 15),
                                                正负=sheet2.cell_value(j, 16),
                                                比赛id=Match.objects.get(id=dir_list[i].split('-')[3]))

                    match_player.save()

                for k in range (8, sheet2.nrows-2): # 替补

                    match_player = Match_player(类型='替补', 主客场=sheet3.cell_value(3, 1), 球员名=sheet2.cell_value(k, 1), 位置=sheet2.cell_value(k, 2), 时间=sheet2.cell_value(k, 3),
                                                投篮=sheet2.cell_value(k, 4), 三分=sheet2.cell_value(k, 5), 罚球=sheet2.cell_value(k, 6),
                                                前场=sheet2.cell_value(k, 7), 后场=sheet2.cell_value(k, 8), 篮板=sheet2.cell_value(k, 9),
                                                助攻=sheet2.cell_value(k, 10), 犯规=sheet2.cell_value(k, 11), 抢断=sheet2.cell_value(k, 12),
                                                失误=sheet2.cell_value(k, 13), 封盖=sheet2.cell_value(k, 14), 得分=sheet2.cell_value(k, 15),
                                                正负=sheet2.cell_value(k, 16),
                                                比赛id=Match.objects.get(id=dir_list[i].split('-')[3]))
                    match_player.save()
                with open("history.txt","a+") as f:
                    f.write("存储："+dir_list[i].split('-')[3]+"\n")
                f.close()

# file_dir2 = "./team_schedule"  # 给定路径
# dir_list = os.listdir(file_dir2)  # 列出文件夹下所有的目录和文件
# new_path = os.path.join(file_dir2, dir_list[0]) # 第二个子文件夹
# file_list = os.listdir(new_path)
#
# for i in range(len(file_list)):
#     real_path1 = os.path.join(new_path, file_list[i])
#     # print(real_path1)
#     workbook1 = xlrd.open_workbook(real_path1)
#     sheet1 = workbook1.sheet_by_index(0)
#
#     for j in range(sheet1.nrows):
#         a = str(sheet1.cell_value(j, 1))
#         b = str(sheet1.cell_value(j, 2))
#         c = str(sheet1.cell_value(j, 4))
#         if len(a.split(' ')) > 3:
#             # print(sheet1.cell_value(j, 1))
#             count_schedule = Schedule.objects.filter(赛季球队=sheet1.cell_value(1, 1), 客队=a.split(' ')[0], 主队=a.split(' ')[4], 日期=c.split(' ')[0])
#             print(count_schedule)
#             if count_schedule == 0: # 判断是否存过赛程
#                 if (b.split(' ')[0] == '-'):  # 比赛前瞻
#                     print(a.split(' ')[0], a.split(' ')[4], sheet1.cell_value(j, 3),
#                           c.split(' ')[0], c.split(' ')[1], sheet1.cell_value(j, 5))
#                     # sched = Schedule(
#                     #     赛季球队=sheet1.cell_value(1, 1),
#                     #     客队=a.split(' ')[0],
#                     #     主队=a.split(' ')[4],
#                     #     客队比分='0',
#                     #     主队比分='0',
#                     #     结果=sheet1.cell_value(j, 3),
#                     #     日期=c.split(' ')[0],
#                     #     北京时间=c.split(' ')[1],
#                     #     类型=sheet1.cell_value(j, 5),
#                     # )
#                     # sched.save()
#                 else:
#                     print(a.split(' ')[0], a.split(' ')[4], b.split(' ')[0], b.split(' ')[2], sheet1.cell_value(j, 3),
#                           c.split(' ')[0], c.split(' ')[1], sheet1.cell_value(j, 5))
#                     print(0)
#                     # sched = Schedule(
#                     #     赛季球队=sheet1.cell_value(1, 1),
#                     #     客队=a.split(' ')[0],
#                     #     主队=a.split(' ')[4],
#                     #     客队比分=b.split(' ')[0],
#                     #     主队比分=b.split(' ')[2],
#                     #     结果=sheet1.cell_value(j, 3),
#                     #     日期=c.split(' ')[0],
#                     #     北京时间=c.split(' ')[1],
#                     #     类型=sheet1.cell_value(j, 5),
#                     # )
#                     # count = Match.objects.filter(主场球队中文名=sched.主队, 客场球队中文名=sched.客队, 日期=sched.日期).count()
#                     # if count!=0:
#                     #     match = Match.objects.get(主场球队中文名=sched.主队, 客场球队中文名=sched.客队, 日期=sched.日期)
#                     #     sched.比赛id=Match(id=match.id)
#                     # sched.save()
#
#             else: # 存过则 判断是否为赛程预测，若存过内容为预测且表中内容为数据统计，则用表内信息覆盖数据库内容
#                 schedule = Schedule.objects.get(赛季球队=sheet1.cell_value(1, 1), 客队=a.split(' ')[0], 主队=a.split(' ')[4], 日期=c.split(' ')[0])
#                 if schedule.类型=='比赛前瞻':
#                     if sheet1.cell_value(j, 5)=='数据统计':
#
#                         print('前瞻->数据统计', a.split(' ')[0], a.split(' ')[4], b.split(' ')[0], b.split(' ')[2], sheet1.cell_value(j, 3),
#                               c.split(' ')[0], c.split(' ')[1], sheet1.cell_value(j, 5))
#                         # schedule.客队比分=b.split(' ')[0]
#                         # schedule.主队比分=b.split(' ')[2]
#                         # schedule.日期=c.split(' ')[0]
#                         # schedule.北京时间=c.split(' ')[1]
#                         # schedule.结果=sheet1.cell_value(j, 3)
#                         # schedule.类型=sheet1.cell_value(j, 5)
#                         # count = Match.objects.filter(主场球队中文名=schedule.主队, 客场球队中文名=schedule.客队, 日期=schedule.日期).count()
#                         # if count != 0:
#                         #     match = Match.objects.get(主场球队中文名=schedule.主队, 客场球队中文名=schedule.客队, 日期=schedule.日期)
#                         #     schedule.比赛id = Match(id=match.id)
#                         # schedule.save()

file_dir2 = "./team_schedule"  # 给定路径
dir_list = os.listdir(file_dir2)  # 列出文件夹下所有的目录和文件
new_path = os.path.join(file_dir2, dir_list[0]) # 第二个子文件夹
file_list = os.listdir(new_path)

for i in range(len(file_list)):
    real_path1 = os.path.join(new_path, file_list[i])
    # print(real_path1)
    workbook1 = xlrd.open_workbook(real_path1)
    sheet1 = workbook1.sheet_by_index(0)

    for j in range(sheet1.nrows):
        a = str(sheet1.cell_value(j, 1))
        b = str(sheet1.cell_value(j, 2))
        c = str(sheet1.cell_value(j, 4))

        if len(a.split(' '))>3:

            if (b.split(' ')[0] == '-'):  # 比赛前瞻
                print(a.split(' ')[0], a.split(' ')[4], sheet1.cell_value(j, 3),
                      c.split(' ')[0], c.split(' ')[1], sheet1.cell_value(j, 5))
                sched = Schedule(
                    英文名=file_list[i].split('.')[0],
                    赛季球队=sheet1.cell_value(1, 1),
                    客队=a.split(' ')[0],
                    主队=a.split(' ')[4],
                    客队比分='0',
                    主队比分='0',
                    结果=sheet1.cell_value(j, 3),
                    日期=c.split(' ')[0],
                    北京时间=c.split(' ')[1],
                    类型=sheet1.cell_value(j, 5),
                )
                sched.save()
            else:
                print(a.split(' ')[0], a.split(' ')[4], b.split(' ')[0], b.split(' ')[2], sheet1.cell_value(j, 3),
                      c.split(' ')[0], c.split(' ')[1], sheet1.cell_value(j, 5))
                print(0)
                sched = Schedule(
                    英文名=file_list[i].split('.')[0],
                    赛季球队=sheet1.cell_value(1, 1),
                    客队=a.split(' ')[0],
                    主队=a.split(' ')[4],
                    客队比分=b.split(' ')[0],
                    主队比分=b.split(' ')[2],
                    结果=sheet1.cell_value(j, 3),
                    日期=c.split(' ')[0],
                    北京时间=c.split(' ')[1],
                    类型=sheet1.cell_value(j, 5),
                )
                count = Match.objects.filter(主场球队中文名=sched.主队, 客场球队中文名=sched.客队, 日期=sched.日期).count()
                if count!=0:
                    match = Match.objects.get(主场球队中文名=sched.主队, 客场球队中文名=sched.客队, 日期=sched.日期)
                    sched.比赛id=Match(id=match.id)

                sched.save()




                #
                # else:
                #     print('')
                #     # print(sheet1.cell_value(j, 1), '数据统计')
        # except:
        #     print("data line")

        # 客场
        # away_summary = Match_teamsummary.objects.get(比赛id=dir_list[i].split('-')[3], 主客场=sheet3.cell_value(2, 1))
        # away_summary.home_away = 1
        # # 主场
        # home_summary = Match_teamsummary.objects.get(比赛id=dir_list[i].split('-')[3], 主客场=sheet3.cell_value(3, 1))
        # home_summary.home_away = 2

            # 比赛id=Match.objects.get(id=dir_list[i].split('-')[3])
            # match_player = Match_player.objects.get(球员名=sheet1.cell_value(j, 1))
            # match_player.类型 = 1
            # match_player.主客场 = 2   # 主场
            #
            # match_player.位置 = sheet1.cell_value(j, 2)
            # match_player.时间 = sheet1.cell_value(j, 3)
            # match_player.投篮 = sheet1.cell_value(j, 4)
            # match_player.三分 = sheet1.cell_value(j, 5)
            # match_player.罚球 = sheet1.cell_value(j, 6)
            # match_player.前场 = sheet1.cell_value(j, 7)
            # match_player.后场 = sheet1.cell_value(j, 8)
            # match_player.篮板 = sheet1.cell_value(j, 9)
            # match_player.助攻 = sheet1.cell_value(j, 10)
            # match_player.犯规 = sheet1.cell_value(j, 11)
            # match_player.抢断 = sheet1.cell_value(j, 12)
            # match_player.失误 = sheet1.cell_value(j, 13)
            # match_player.封盖 = sheet1.cell_value(j, 14)
            # match_player.得分 = sheet1.cell_value(j, 15)
            # match_player.正负 = sheet1.cell_value(j, 16)
            # match_player.比赛id = dir_list[i].split('-')[3]
            # match_player.save()
            #

            # for k in range (9,sheet1.nrows-2):
            #     print(sheet1.cell_value(k, 1))
            #     k = k+1
        # print(real_path1, real_path2, real_path3)
        # print(dir_list[i].split('-')[0] + '-' + dir_list[i].split('-')[1] + '-' + dir_list[i].split('-')[2])
        #
    # print(real_path)
    # image = open(real_path,'rb')
    # image = File(image)
    # player = Player.objects.get(序号=dir_list[i].split('-')[1])
    # player.头像 = image
    # player.save()
    # with open(real_path,'rb') as f:
    #     file_text = json.load(f)
    # if j==0 :
    #     file_text['类型']=0
    # if j==1 :
    #     file_text['类型']=1
    # if j==2 :
    #     file_text['类型']=2
    #
    # file_text['序号']=dir_list[i].split('-')[1]
    # content = JSONRenderer().render(file_text)
    # stream = BytesIO(content)
    # data = JSONParser().parse(stream)
    # serializer = CareerSerializer(data=data)
    # if serializer.is_valid():
    #     serializer.save()
    # else:
    #     print(serializer.errors)



# file_dir2="./player_profile_json"
# dir_list = os.listdir(file_dir2) # 列出文件夹下所有的目录和文件
# i=0
# for i in range(len(dir_list)):
#     new_path = os.path.join(file_dir2,dir_list[i])
#     # print(dir_list[i])
#
#     with open(new_path,'rb') as f:
#         file_text = json.load(f)
#
#     info = file_text[dir_list[i].split('.')[0]]
#     for item in info:
#         player = Player.objects.get(序号=item['player_id'])
#         #print(dir_list[i].split('.')[0])
#         player.球队名 = Team.objects.get(球队名=dir_list[i].split('.')[0])
#         player.save()
#     i = i+1
# print(i)





